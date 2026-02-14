import openpyxl as xl
from openpyxl.chart import BarChart,Reference

wb=xl.load_workbook('transactions.xlsx')
sheet=wb['Sheet1']
cell=sheet['a1']
sheet.cell(1,1)
print(cell.value)
len=sheet.max_row()
for row in range(2,len+1):
    sheet.cell(row,3)
    print(cell.value)
    corrected_value=cell.value*0.9
    corrected_value_cell=sheet.cell(row,4)
    corrected_value_cell.value=corrected_value

values=Reference(sheet,min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)

chart=BarChart()
chart.add_data(values)
sheet.add_chart(chart,'e2')

wb.save('transactions2.xlsx')
